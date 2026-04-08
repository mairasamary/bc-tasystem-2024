from django.views import View
from django.shortcuts import render, redirect
from .models import Course
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from openpyxl import load_workbook
from users.models import CustomUser as User
from random import randint
from django.views.generic import ListView, DetailView
from django.contrib import messages
from users.instructor_data import instructors

from random import randint


class UploadView(LoginRequiredMixin, UserPassesTestMixin, View):
    def get(self, request):
        return render(request, "upload.html")

    def post(self, request):
        excel_file = request.FILES.get("excel_file")
        if excel_file:
            self.process_excel_file(excel_file)
        messages.success(self.request, "Successfully Uploaded Excel File")
        return render(request, "upload.html")

    def process_excel_file(self, excel_file):
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        header_map = self.build_header_map(header_row)

        for row in rows:
            row_data = self.parse_row_data(row, header_map)
            if not row_data.get("term"):
                continue
            instructor = self.get_or_create_instructor(row_data)
            self.create_course(row_data, instructor)

    def build_header_map(self, header_row):
        if not header_row:
            return {}
        mapping = {}
        for idx, header in enumerate(header_row):
            if not header:
                continue
            normalized = str(header).strip().lower().replace(" ", "")
            mapping[normalized] = idx
        return mapping

    def _get_value(self, row, header_map, key, fallback_index):
        idx = header_map.get(key, fallback_index)
        if idx is None:
            return None
        if idx >= len(row):
            return None
        return row[idx]

    def _to_int(self, value, default=0):
        try:
            if value is None or value == "":
                return default
            return int(float(value))
        except (TypeError, ValueError):
            return default

    def parse_row_data(self, row, header_map):
        # Supports both new and legacy Excel layouts.
        return {
            "term": self._get_value(row, header_map, "term", 0),
            "class_type": self._get_value(row, header_map, "type", 1),
            "course": self._get_value(row, header_map, "course", 2),
            "section": self._get_value(row, header_map, "section", 3),
            "course_title": self._get_value(row, header_map, "coursetitle", 4),
            "instructors": self._get_value(row, header_map, "instructors", 5),
            "room_name": self._get_value(row, header_map, "roomname", 6),
            "timeslot": self._get_value(row, header_map, "timeslot", 7),
            "max_enroll": self._to_int(self._get_value(row, header_map, "maxenroll", 8), 0),
            "room_size": self._to_int(self._get_value(row, header_map, "roomsize", 9), 0),
        }

    def get_email(self, first_name, last_name):
        return (
            f"{first_name}.{last_name}@bc.edu"
            if f"{last_name}, {first_name}" not in instructors
            else instructors[f"{last_name}, {first_name}"]["Short Email"]
        )

    def get_or_create_instructor(self, row_data):
        instructor_value = row_data.get("instructors")
        if not instructor_value:
            instructor_value = "Unknown"
        try:
            instructor_first_name = instructor_value.split(",")[1].strip()
            instructor_last_name = instructor_value.split(",")[0].strip()
        except:
            instructor_first_name = str(instructor_value)
            instructor_last_name = ""

        email = self.get_email(instructor_first_name, instructor_last_name)

        instructor, created = User.objects.get_or_create(
            email=email,
            defaults={
                "professor": True,
                "eagleid": self.generate_eagleid(),
                "first_name": instructor_first_name,
                "last_name": instructor_last_name,
            },
        )

        if created:
            instructor.set_password("password")
            instructor.save()

        return instructor

    def create_course(self, row_data, instructor):
        excluded_lectures = [
            "Computer Science I",
            "Computer Science II",
            "Computer Organization and Lab",
        ]

        course_title = str(row_data.get("course_title") or "").strip()
        class_type = str(row_data.get("class_type") or "").strip()
        max_enroll = self._to_int(row_data.get("max_enroll"), 0)

        if course_title in excluded_lectures and class_type == "Lecture":
            return

        num_tas = 1 if class_type in ("Discussion", "Lab") else max_enroll // 20

        num_tas = 1 if num_tas == 0 else num_tas

        new_class = Course.objects.create(
            term=row_data.get("term"),
            class_type=class_type,
            course=row_data.get("course"),
            section=row_data.get("section"),
            course_title=course_title,
            instructor_first_name=instructor.first_name,
            instructor_last_name=instructor.last_name,
            room_name=str(row_data.get("room_name") or ""),
            timeslot=str(row_data.get("timeslot") or ""),
            max_enroll=max_enroll,
            room_size=self._to_int(row_data.get("room_size"), 0),
            num_tas=num_tas,
            description=course_title,  # TODO: Add description
            professor=instructor,
        )
        instructor.courses.add(new_class)

    def generate_eagleid(self):
        eagleid = randint(10000000, 99999999)
        while User.objects.filter(eagleid=eagleid).exists():
            eagleid = randint(10000000, 99999999)
        return eagleid

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser


class CloseView(LoginRequiredMixin, UserPassesTestMixin, View):
    def post(self, request):
        self.archive_courses()
        messages.success(self.request, "Successfully Closed Courses")
        return redirect("courses:manage-course")

    def archive_courses(self):
        current_courses = Course.objects.all()
        for course in current_courses:
            course.is_active = False
            course.save()

    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_superuser


class ListView(LoginRequiredMixin, ListView):
    model = Course
    template_name = "list.html"
    ordering = ["course"]
    context_object_name = "course_data"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["professors"] = User.objects.filter(professor=True)
        return context

    def get_queryset(self):
        user = self.request.user
        professor_id = self.request.GET.get("professor_id", None)

        if professor_id:
            return Course.objects.filter(professor__id=professor_id, is_active=True)

        if user.is_student or user.is_superuser:
            return Course.objects.filter(is_active=True)

        return Course.objects.filter(professor=user, is_active=True)


class CourseDetailView(LoginRequiredMixin, DetailView):
    model = Course
    template_name = "course_detail.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["course"] = Course.objects.get(pk=self.kwargs.get("pk"))
        context["is_professor"] = self.is_professor()
        context["at_max_apps"] = self.at_max_apps()
        context["has_applied"] = self.has_applied()
        context["is_ta"] = self.is_ta()
        return context

    def is_professor(self):
        return self.request.user.is_professor

    def at_max_apps(self):
        return self.request.user.reached_max_applications()

    def has_applied(self):
        return self.request.user.already_applied_to_course(self.get_object())

    def is_ta(self):
        return self.request.user.is_ta()
