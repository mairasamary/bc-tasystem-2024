
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.db.models import Count, F, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from datetime import date
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from courses.models import Course, CourseQuestion
from evaluations.models import TAEvaluation
from applications.models import Application, ApplicationStatus, ApplicationAnswer
from applications.forms import ApplicationForm
from offers.models import Offer, OfferStatus
from users.forms import StudentEmploymentOnboardingForm
from users.models import StudentProfile, PastCourse
from main.models import Notification
from main.constants import BC_STUDENT_EMPLOYMENT_NEW_HIRES_URL
from main.notifications import create_notification
from main.ta_hiring import send_student_ta_acceptance_onboarding_email
from main.utils import app_site_absolute_url, send_notification_email

User = get_user_model()

# Shared ordering for the applications table and professor/superuser prev/next on detail pages
APPLICATIONS_SORT_ORDER = (
    "student__last_name",
    "student__first_name",
    "course__course",
    "course__section",
    "id",
)


def home(request):
    return render(request, "home.html")


def contributors(request):
    return render(request, "contributors.html")


@login_required
def profile_welcome(request):
    """One-time welcome for new students; POST acknowledges and sends them to profile."""
    if not request.user.student_needs_profile_welcome():
        return redirect("dashboard")
    if request.method == "POST":
        bc_worker = bool(request.POST.get("bc_student_worker"))
        profile, _ = StudentProfile.objects.get_or_create(user=request.user)
        profile.bc_student_worker = bc_worker
        profile.save(update_fields=["bc_student_worker"])
        request.user.profile_welcome_acknowledged = True
        request.user.save(update_fields=["profile_welcome_acknowledged"])
        return redirect("student_profile")
    return render(request, "profile_welcome.html")


@login_required
def notifications_page(request):
    notifications = Notification.objects.filter(user=request.user).order_by("-created_at")
    return render(request, "notifications.html", {"notifications": notifications})


@login_required
def open_notification(request, notification_id):
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    if not notification.is_read:
        notification.is_read = True
        notification.save(update_fields=["is_read"])
    return redirect(notification.target_url or "/")


@login_required
def dismiss_notification(request, notification_id):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Method not allowed"}, status=405)
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    notification.delete()
    return JsonResponse({"ok": True})


@login_required
def clear_notifications(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Method not allowed"}, status=405)
    Notification.objects.filter(user=request.user).delete()
    return JsonResponse({"ok": True})


@login_required
def admin_dashboard_v2(request):
    # Admin: full system overview (stats + recent apps/offers)
    if request.user.is_superuser:
        pending_apps_count = Application.objects.filter(status=ApplicationStatus.PENDING.value).count()
        total_offers_count = Offer.objects.count()
        active_courses_count = Course.objects.filter(status=True).count()
        total_users_count = User.objects.count()
        pending_apps = Application.objects.filter(status=ApplicationStatus.PENDING.value).select_related('student', 'course').order_by('-id')[:5]
        recent_offers = Offer.objects.all().select_related('recipient', 'course').order_by('-created_at')[:5]
        context = {
            'pending_apps_count': pending_apps_count,
            'total_offers_count': total_offers_count,
            'active_courses_count': active_courses_count,
            'total_users_count': total_users_count,
            'pending_apps': pending_apps,
            'recent_offers': recent_offers,
        }
        return render(request, 'dashboard.html', context)

    # Professor: 4 stat cards + Course Staffing Overview table
    if request.user.is_professor:
        my_active_courses_count = Course.objects.filter(
            professor=request.user, status=True
        ).count()
        understaffed_count = (
            Course.objects.filter(professor=request.user, status=True)
            .annotate(tas_count=Count('current_tas', distinct=True))
            .filter(tas_count__lt=F('num_tas'))
            .count()
        )
        pending_applications_count = Application.objects.filter(
            course__professor=request.user,
            status=ApplicationStatus.PENDING.value,
        ).count()
        pending_offers_count = Offer.objects.filter(
            sender=request.user, status=OfferStatus.PENDING.value
        ).count()
        staffing_overview = (
            Course.objects.filter(professor=request.user)
            .annotate(
                tas_count=Count('current_tas', distinct=True),
                pending_apps_count=Count(
                    'applications',
                    filter=Q(applications__status=ApplicationStatus.PENDING.value),
                    distinct=True,
                ),
                pending_offers_count=Count(
                    'offer_course',
                    filter=Q(
                        offer_course__status=OfferStatus.PENDING.value,
                        offer_course__sender=request.user,
                    ),
                    distinct=True,
                ),
            )
            .order_by('-term', 'course')
        )
        context = {
            'my_active_courses_count': my_active_courses_count,
            'understaffed_count': understaffed_count,
            'pending_applications_count': pending_applications_count,
            'pending_offers_count': pending_offers_count,
            'staffing_overview': staffing_overview,
        }
        return render(request, 'professor_dashboard.html', context)

    # Student Dashboard: stats for minimal, action-oriented UI
    my_apps = Application.objects.filter(student=request.user).select_related('course').order_by('-id')
    my_offers = Offer.objects.filter(recipient=request.user).select_related('course', 'sender').order_by('-created_at')
    assigned_course = request.user.course_working_for.first()
    current_term = (assigned_course.term or "").strip() if assigned_course else ""
    if not current_term:
        t = Course.objects.order_by('-term').values_list('term', flat=True).first()
        current_term = (t or "").strip()
    counted_statuses = [
        ApplicationStatus.PENDING.value,
        ApplicationStatus.ACCEPTED.value,
        ApplicationStatus.CONFIRMED.value,
    ]
    applications_active_this_term = (
        Application.objects.filter(
            student=request.user,
            course__term__iexact=current_term,
            status__in=counted_statuses,
        ).count()
        if current_term
        else 0
    )
    offers_awaiting_response = Offer.objects.filter(
        recipient=request.user, status=OfferStatus.PENDING.value
    ).count()
    has_accepted_offer_this_term = Offer.objects.filter(
        recipient=request.user, status=OfferStatus.ACCEPTED.value
    ).exists()
    assigned_course_name = assigned_course.course if assigned_course else None
    context = {
        'my_apps': my_apps,
        'my_offers': my_offers,
        'applications_active_this_term': applications_active_this_term,
        'offers_awaiting_response': offers_awaiting_response,
        'has_accepted_offer_this_term': has_accepted_offer_this_term,
        'assigned_course_name': assigned_course_name,
    }
    return render(request, 'student_dashboard.html', context)


@login_required
def employment_onboarding_checklist(request):
    if request.user.is_professor and not request.user.is_superuser:
        messages.error(request, "This page is for students with a TA assignment.")
        return redirect("dashboard")
    if not request.user.has_ta_assignment:
        messages.error(
            request,
            "This checklist is available after you have accepted a TA assignment.",
        )
        return redirect("dashboard")
    profile, _ = StudentProfile.objects.get_or_create(user=request.user)

    # Past TAs and BC student workers do not need to re-complete these checklist forms.
    if profile.onboarding_complete:
        messages.info(request, "Your student employment onboarding is already complete. No checklist needed.")
        return redirect("dashboard")

    if request.method == "POST":
        form = StudentEmploymentOnboardingForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, "Your onboarding checklist has been saved.")
            return redirect("employment_onboarding")
    else:
        form = StudentEmploymentOnboardingForm(instance=profile)
    return render(
        request,
        "employment_onboarding_checklist.html",
        {
            "form": form,
            "bc_new_hires_url": BC_STUDENT_EMPLOYMENT_NEW_HIRES_URL,
        },
    )


@login_required
def applications_list_v2(request):
    qs = Application.objects.select_related("student", "course")
    if request.user.is_superuser:
        apps = qs
    elif request.user.is_professor:
        apps = qs.filter(course__professor=request.user)
    else:
        apps = qs.filter(student=request.user)
    apps = apps.order_by(*APPLICATIONS_SORT_ORDER)
    return render(request, "applications.html", {"apps": apps})

@login_required
def offers_list_v2(request):
    # Professors see only offers they sent (even if also superuser)
    if request.user.is_professor:
        offers = Offer.objects.filter(sender=request.user).select_related('recipient', 'course').order_by('-created_at')
    elif request.user.is_superuser:
        offers = Offer.objects.select_related('recipient', 'course', 'sender').order_by('-created_at')
    else:
        offers = Offer.objects.filter(recipient=request.user).select_related('recipient', 'course').order_by('-created_at')
    return render(request, 'offers.html', {'offers': offers})

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse, FileResponse
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from applications.models import Application, ApplicationStatus
from courses.forms import CourseForm
from courses.views import UploadView

PER_PAGE_CHOICES = [10, 20, 50]

@login_required
def courses_list_v2(request):
    query = request.GET.get('q')
    term_filter = (request.GET.get('term') or '').strip()
    professor_id = request.GET.get('professor')
    status_filter = request.GET.get('status', '')
    class_type_filter = request.GET.get('class_type', '')
    course_level = request.GET.get('course_level', '')
    per_page_param = request.GET.get('per_page', '10')
    show_all = request.GET.get('show_all', '')
    courses = Course.objects.all().order_by('course')

    # Professors see only their own courses by default
    professor_my_courses = False
    if request.user.is_professor and not request.user.is_superuser:
        if not show_all:
            courses = courses.filter(professor=request.user)
            professor_my_courses = True

    if query:
        courses = courses.filter(
            Q(course__icontains=query) |
            Q(course_title__icontains=query) |
            Q(instructor_first_name__icontains=query) |
            Q(instructor_last_name__icontains=query)
        )
    if term_filter:
        courses = courses.filter(term__iexact=term_filter)
    if professor_id:
        courses = courses.filter(professor_id=professor_id)
    if status_filter == 'active':
        courses = courses.filter(status=True)
    elif status_filter == 'closed':
        courses = courses.filter(status=False)
    if class_type_filter:
        courses = courses.filter(class_type=class_type_filter)
    if course_level and course_level in ('1', '2', '3', '4', '5'):
        courses = courses.filter(course__iregex=r'^\D*' + course_level + r'\d{3}')

    # Admin-only stats and drill-down (on filtered queryset, before pagination)
    course_stats = None
    courses_needing_tas = None
    understaffed_by_professor = None
    staffing_filter_urls = None
    staffing_filter = (request.GET.get('staffing') or '').strip()

    if request.user.is_superuser:
        stats_qs = courses.annotate(tas_count=Count('current_tas', distinct=True))
        agg = stats_qs.aggregate(
            total_slots=Sum('num_tas'),
            filled_slots=Sum('tas_count'),
        )
        total_slots = agg['total_slots'] or 0
        filled_slots = agg['filled_slots'] or 0
        understaffed = stats_qs.filter(tas_count__lt=F('num_tas')).count()
        no_tas_yet = stats_qs.filter(tas_count=0, num_tas__gt=0).count()
        fully_staffed = stats_qs.filter(tas_count__gte=F('num_tas'), num_tas__gt=0).count()
        course_stats = {
            'total_courses': courses.count(),
            'understaffed': understaffed,
            'no_tas_yet': no_tas_yet,
            'partially_filled': understaffed - no_tas_yet,
            'fully_staffed': fully_staffed,
            'total_ta_slots': total_slots,
            'filled_ta_slots': filled_slots,
        }
        # Drill-down: courses needing TAs (for compact table)
        needing = stats_qs.filter(tas_count__lt=F('num_tas')).select_related('professor').order_by('course')[:25]
        courses_needing_tas = [
            {
                'id': c.id,
                'course': c.course,
                'section': c.section,
                'course_title': c.course_title,
                'professor_name': c.professor.get_full_name() or (c.professor.email if c.professor else '') or f'{c.instructor_last_name}, {c.instructor_first_name}'.strip(', ') or '—',
                'ta_count': c.tas_count,
                'num_tas': c.num_tas,
            }
            for c in needing
        ]
        # Understaffed count by professor (for chart)
        by_prof_ids = list(
            stats_qs.filter(tas_count__lt=F('num_tas'))
            .values('professor_id')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        if by_prof_ids:
            ids = [x['professor_id'] for x in by_prof_ids if x['professor_id']]
            users = {u.id: u.get_full_name() or u.email or '—' for u in User.objects.filter(id__in=ids)} if ids else {}
            understaffed_by_professor = [
                {'name': users.get(x['professor_id'], '—') if x['professor_id'] else '—', 'count': x['count']}
                for x in by_prof_ids
            ]
        else:
            understaffed_by_professor = []
        # Apply staffing filter to table (narrow what we paginate)
        if staffing_filter == 'no_tas':
            courses = stats_qs.filter(tas_count=0, num_tas__gt=0).order_by('course')
        elif staffing_filter == 'understaffed':
            courses = stats_qs.filter(tas_count__lt=F('num_tas')).order_by('course')
        elif staffing_filter == 'fully_staffed':
            courses = stats_qs.filter(tas_count__gte=F('num_tas'), num_tas__gt=0).order_by('course')
        # URLs for chart segment clicks (preserve other GET params)
        _get = request.GET.copy()
        _get.pop('page', None)
        def url_with_staffing(val):
            c = _get.copy()
            if val:
                c['staffing'] = val
            else:
                c.pop('staffing', None)
            return request.path + ('?' + c.urlencode() if c else '')
        staffing_filter_urls = {
            'no_tas': url_with_staffing('no_tas'),
            'understaffed': url_with_staffing('understaffed'),
            'fully_staffed': url_with_staffing('fully_staffed'),
            'clear': url_with_staffing(''),
        }

    if per_page_param == 'all':
        per_page_size = 9999
        per_page = 'all'
    else:
        try:
            n = int(per_page_param)
            per_page_size = n if n in PER_PAGE_CHOICES else 10
            per_page = str(per_page_size)
        except (ValueError, TypeError):
            per_page_size = 10
            per_page = '10'

    paginator = Paginator(courses, per_page_size)
    page_number = request.GET.get('page', 1)
    page = paginator.get_page(page_number)
    courses = page.object_list

    professors = User.objects.filter(professor=True).order_by('last_name', 'first_name')
    terms = list(Course.objects.values_list('term', flat=True).distinct().order_by('-term'))

    if not request.user.is_professor:
        applied_course_ids = Application.objects.filter(student=request.user).values_list('course_id', flat=True)
        # For disabling Apply entrypoints: limit 5 per term or already assigned this term
        assigned_course = request.user.course_working_for.first()
        current_term = (assigned_course.term or "").strip() if assigned_course else ""
        if not current_term:
            t = Course.objects.order_by('-term').values_list('term', flat=True).first()
            current_term = (t or "").strip()
        counted = [
            ApplicationStatus.PENDING.value,
            ApplicationStatus.ACCEPTED.value,
            ApplicationStatus.CONFIRMED.value,
        ]
        apps_this_term = (
            Application.objects.filter(
                student=request.user,
                course__term__iexact=current_term,
                status__in=counted,
            ).count()
            if current_term
            else 0
        )
        # One TA position total: cannot apply after accepting / being assigned as TA
        student_can_apply = not (apps_this_term >= 5 or request.user.is_ta())
    else:
        applied_course_ids = []
        student_can_apply = True  # professors don't use Apply

    get_copy = request.GET.copy()
    if 'page' in get_copy:
        del get_copy['page']
    query_string = get_copy.urlencode()

    return render(request, 'courses.html', {
        'courses': courses,
        'applied_course_ids': applied_course_ids,
        'professors': professors,
        'terms': terms,
        'page': page,
        'paginator': paginator,
        'query_string': query_string,
        'per_page': per_page,
        'student_can_apply': student_can_apply,
        'course_stats': course_stats,
        'courses_needing_tas': courses_needing_tas,
        'understaffed_by_professor': understaffed_by_professor or [],
        'staffing_filter_urls': staffing_filter_urls or {},
        'staffing_filter': staffing_filter,
        'professor_my_courses': professor_my_courses,
    })

@login_required
def create_course_v2(request):
    if not request.user.is_superuser:
        messages.error(request, "Only admins can create courses.")
        return redirect('courses')

    if request.method == 'POST':
        form = CourseForm(request.POST)
        if form.is_valid():
            course = form.save(commit=False)
            if course.professor:
                course.instructor_first_name = course.professor.first_name
                course.instructor_last_name = course.professor.last_name
            else:
                course.instructor_first_name = ''
                course.instructor_last_name = ''
            course.save()
            messages.success(request, f"Course {course.course} - {course.course_title} created successfully.")
            return redirect('courses')
    else:
        form = CourseForm()

    return render(request, 'create_course.html', {'form': form})


@login_required
def course_overview_v2(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    context = {'course': course}
    if not request.user.is_professor:
        applied_course_ids = list(
            Application.objects.filter(student=request.user).values_list('course_id', flat=True)
        )
        context['applied_course_ids'] = applied_course_ids
        assigned_course = request.user.course_working_for.first()
        current_term = (assigned_course.term or "").strip() if assigned_course else ""
        if not current_term:
            t = Course.objects.order_by('-term').values_list('term', flat=True).first()
            current_term = (t or "").strip()
        counted = [
            ApplicationStatus.PENDING.value,
            ApplicationStatus.ACCEPTED.value,
            ApplicationStatus.CONFIRMED.value,
        ]
        apps_this_term = (
            Application.objects.filter(
                student=request.user,
                course__term__iexact=current_term,
                status__in=counted,
            ).count()
            if current_term
            else 0
        )
        context['student_can_apply'] = not (apps_this_term >= 5 or request.user.is_ta())
    else:
        context['applied_course_ids'] = []
        context['student_can_apply'] = False
    return render(request, 'course_overview.html', context)


@login_required
def add_course_question_v2(request, course_id):
    if request.method != 'POST':
        return redirect('course_overview', course_id=course_id)
    course = get_object_or_404(Course, id=course_id)
    if not (request.user.is_superuser or request.user == course.professor):
        messages.error(request, "Only the course professor can manage questions.")
        return redirect('course_overview', course_id=course_id)
    if course.custom_questions.count() >= 5:
        messages.error(request, "Maximum of 5 custom questions per course.")
        return redirect('course_overview', course_id=course_id)
    question_text = (request.POST.get('question_text') or '').strip()
    if not question_text:
        messages.error(request, "Question text cannot be empty.")
        return redirect('course_overview', course_id=course_id)
    if len(question_text) > 300:
        messages.error(request, "Question text must be 300 characters or fewer.")
        return redirect('course_overview', course_id=course_id)
    is_required = request.POST.get('is_required') == 'on'
    next_order = (course.custom_questions.order_by('-order').values_list('order', flat=True).first() or 0) + 1
    CourseQuestion.objects.create(
        course=course,
        question_text=question_text,
        is_required=is_required,
        order=next_order,
    )
    messages.success(request, "Question added.")
    return redirect('course_overview', course_id=course_id)


@login_required
def delete_course_question_v2(request, course_id, question_id):
    if request.method != 'POST':
        return redirect('course_overview', course_id=course_id)
    course = get_object_or_404(Course, id=course_id)
    if not (request.user.is_superuser or request.user == course.professor):
        messages.error(request, "Only the course professor can manage questions.")
        return redirect('course_overview', course_id=course_id)
    question = get_object_or_404(CourseQuestion, id=question_id, course=course)
    if question.has_answers:
        messages.error(request, "Cannot delete a question that already has student answers.")
        return redirect('course_overview', course_id=course_id)
    question.delete()
    messages.success(request, "Question deleted.")
    return redirect('course_overview', course_id=course_id)


@login_required
def edit_course_v2(request, course_id):
    if not request.user.is_superuser:
        return redirect('course_overview', course_id=course_id)
    course = get_object_or_404(Course, id=course_id)
    if request.method == 'POST':
        form = CourseForm(request.POST, instance=course)
        if form.is_valid():
            form.save()
            # Courses only close when full (TA capacity) or admin closes semester; enforce full => closed on edit
            if course.current_tas.count() >= course.num_tas:
                course.status = False
                course.save(update_fields=['status'])
            messages.success(request, f"Course {course.course} - {course.course_title} updated successfully.")
            return redirect('courses')
    else:
        form = CourseForm(instance=course)
    return render(request, 'edit_course.html', {'form': form, 'course': course})


@login_required
def upload_courses_v2(request):
    if not request.user.is_superuser:
        messages.error(request, "Only admins can upload courses.")
        return redirect('courses')

    if request.method == 'POST':
        excel_file = request.FILES.get("excel_file")
        if excel_file:
            upload_view = UploadView()
            upload_view.request = request
            upload_view.process_excel_file(excel_file)
            messages.success(request, "Successfully uploaded courses from Excel file.")
        else:
            messages.error(request, "Please select an Excel file to upload.")
        return redirect('courses')

    return render(request, 'upload_courses.html')

@login_required
def close_semester_v2(request):
    if not request.user.is_superuser:
        messages.error(request, "Only admins can close the semester.")
        return redirect('courses')

    if request.method == 'POST':
        # "Close semester" is implemented as a system-wide course close:
        # - `status=False` blocks student applications (`apply_to_course_v2` checks `course.status`)
        # - `is_active=False` hides courses from professor/student course pickers
        #
        # We additionally notify professors so they can leave optional TA evaluations.
        open_courses = list(
            Course.objects.filter(status=True)
            .select_related("professor")
            .prefetch_related("current_tas")
        )

        # Close all courses for the semester (the other way courses close; the other is full TA capacity)
        Course.objects.all().update(is_active=False, status=False)

        if not open_courses:
            messages.success(request, "Successfully closed all courses for the semester.")
            return redirect('courses')

        # Map: professor_id -> {prof, pairs[(ta_id, course_id)]}
        professor_to_pairs = {}
        ta_ids = set()
        closed_course_ids = {c.id for c in open_courses}

        for course in open_courses:
            professor = course.professor
            if not professor:
                continue
            bucket = professor_to_pairs.setdefault(professor.id, {"prof": professor, "pairs": []})
            for ta in course.current_tas.all():
                bucket["pairs"].append((ta.id, course.id))
                ta_ids.add(ta.id)

        # Only notify professors who actually have at least one TA in a course being closed.
        if not professor_to_pairs:
            messages.success(request, "Successfully closed all courses for the semester.")
            return redirect('courses')

        existing_evals = set(
            TAEvaluation.objects.filter(
                reviewer_id__in=list(professor_to_pairs.keys()),
                ta_id__in=list(ta_ids),
                course_id__in=list(closed_course_ids),
            ).values_list("reviewer_id", "ta_id", "course_id")
        )

        for professor_id, bucket in professor_to_pairs.items():
            professor = bucket["prof"]
            pairs = bucket["pairs"]
            if not pairs:
                continue

            # Prefer a link to a missing evaluation; if all are already reviewed, fall back to the first pair.
            missing_pair = next(
                ((ta_id, course_id) for (ta_id, course_id) in pairs if (professor_id, ta_id, course_id) not in existing_evals),
                None,
            )
            link_ta_id, link_course_id = missing_pair or pairs[0]

            target_url = (
                f"{reverse('evaluations:create')}?ta_id={link_ta_id}&course_id={link_course_id}"
            )

            create_notification(
                user=professor,
                title="TA evaluations requested (optional written feedback)",
                target_url=target_url,
            )

            if professor.email:
                send_notification_email(
                    subject="TA Evaluation Request",
                    recipients=professor.email,
                    message_lines=[
                        f"Dear {professor.get_full_name()},",
                        "The semester has been closed. Please leave TA evaluations for your TAs below.",
                        f"Review link: {app_site_absolute_url(target_url)}",
                        "Written feedback is optional.",
                    ],
                )

        messages.success(request, "Successfully closed all courses for the semester and prompted professors to complete TA evaluations.")

    return redirect('courses')


BC_NEW_HIRES_URL = "https://www.bc.edu/content/bc-web/offices/student-services/student-employment/new-hires.html"
CS_TA_GUIDELINES_CONTRACT_URL = "https://docs.google.com/forms/d/e/1FAIpQLSfcb4M_9cD8CW_nz8eQrhEKiinBjpobpTMkTHe-TTBDODv4iQ/viewform"

ONBOARDING_STEPS = [
    ("onboarding_done_required_form", "CS TA Guidelines Contract"),
    ("onboarding_done_i9", "I-9 (new hires only)"),
    ("onboarding_done_payroll_statement", "SSN process (international new hires)"),
    ("onboarding_done_w4", "Kronos timecard visible"),
    ("onboarding_done_m4", "PeopleSoft tax forms"),
    ("onboarding_done_direct_deposit", "PeopleSoft direct deposit"),
]


def _build_onboarding_rows():
    """Return a list of dicts — one per TA — with their onboarding status."""
    tas = User.objects.filter(course_working_for__isnull=False).distinct().order_by('last_name', 'first_name')
    rows = []
    for ta in tas:
        try:
            profile = ta.student_profile
        except Exception:
            profile = None
        if profile:
            profile_onboarding_complete = bool(profile.onboarding_complete)
        else:
            profile_onboarding_complete = bool(getattr(ta, "has_been_ta_before", lambda: False)())
        steps = []
        completed = 0
        for field, label in ONBOARDING_STEPS:
            done = profile_onboarding_complete or (bool(getattr(profile, field, False)) if profile else False)
            if done:
                completed += 1
            steps.append({'label': label, 'done': done})
        course = ta.course_working_for.first()
        rows.append({
            'user': ta,
            'course': course,
            'steps': steps,
            'completed': completed,
            'total': len(ONBOARDING_STEPS),
            'is_complete': profile_onboarding_complete or completed == len(ONBOARDING_STEPS),
        })
    return rows


@login_required
def onboarding_status_v2(request):
    if not request.user.is_superuser:
        messages.error(request, "Only admins can view onboarding status.")
        return redirect('dashboard')
    rows = _build_onboarding_rows()
    complete_count = sum(1 for r in rows if r['is_complete'])
    incomplete_count = len(rows) - complete_count
    return render(request, 'onboarding_status.html', {
        'rows': rows,
        'step_labels': [label for _, label in ONBOARDING_STEPS],
        'complete_count': complete_count,
        'incomplete_count': incomplete_count,
        'total_count': len(rows),
    })


@login_required
def export_onboarding_status(request):
    if not request.user.is_superuser:
        messages.error(request, "Only admins can export onboarding status.")
        return redirect('dashboard')

    rows = _build_onboarding_rows()

    wb = Workbook()
    ws = wb.active
    ws.title = "Onboarding Status"

    headers = ['Name', 'Email', 'Course'] + [label for _, label in ONBOARDING_STEPS] + ['Completed', 'Status']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        data = [
            row['user'].get_full_name(),
            row['user'].email,
            row['course'].course if row['course'] else '—',
        ]
        for step in row['steps']:
            data.append('Yes' if step['done'] else 'No')
        data.append(f"{row['completed']}/{row['total']}")
        data.append('Complete' if row['is_complete'] else 'Incomplete')
        ws.append(data)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="onboarding_status.xlsx"'
    wb.save(response)
    return response


@login_required
def send_onboarding_reminders(request):
    if request.method != 'POST':
        return redirect('onboarding_status')
    if not request.user.is_superuser:
        messages.error(request, "Only admins can send onboarding reminders.")
        return redirect('dashboard')

    rows = _build_onboarding_rows()
    incomplete = [r for r in rows if not r['is_complete']]
    sent = 0
    for row in incomplete:
        ta = row['user']
        if not ta.email:
            continue
        missing = [s['label'] for s in row['steps'] if not s['done']]
        missing_list = '\n'.join(f'  • {m}' for m in missing)
        send_notification_email(
            subject="Action Required: Complete Your TA Onboarding Forms",
            recipients=ta.email,
            message_lines=[
                f"Dear {ta.get_full_name()},",
                f"You have been assigned as a TA for {row['course'].course if row['course'] else 'your course'}, but your BC student employment onboarding is not yet complete.",
                f"You are still missing the following checklist item(s):\n{missing_list}",
                "Step 1 for all TAs: review and sign the CS Teaching Assistant Guidelines Contract:",
                CS_TA_GUIDELINES_CONTRACT_URL,
                "New hires who still need an I-9 must complete it in person at Student Employment in Lyons with original documents.",
                "International new hires who need a Social Security number should follow BC's international student instructions on the same page.",
                f"BC Student Employment New Hires instructions:\n{BC_NEW_HIRES_URL}",
                "Once your paperwork is processed and your timecard appears in Kronos, complete direct deposit and tax forms in PeopleSoft HR.",
                "If an item does not apply to you (for example, you already have an I-9 on file), you can still mark your checklist accordingly in TA Connect.",
            ],
        )
        sent += 1

    messages.success(request, f"Reminder sent to {sent} student{'s' if sent != 1 else ''} with incomplete onboarding.")
    return redirect('onboarding_status')


def _get_export_queryset(request):
    """Apply same filters as courses list; used for export."""
    query = request.GET.get('q')
    term_filter = (request.GET.get('term') or '').strip()
    professor_id = request.GET.get('professor')
    status_filter = request.GET.get('status', '')
    class_type_filter = request.GET.get('class_type', '')
    course_level = request.GET.get('course_level', '')
    courses = Course.objects.all().order_by('course')
    if query:
        courses = courses.filter(
            Q(course__icontains=query) |
            Q(course_title__icontains=query) |
            Q(instructor_first_name__icontains=query) |
            Q(instructor_last_name__icontains=query)
        )
    if term_filter:
        courses = courses.filter(term__iexact=term_filter)
    if professor_id:
        courses = courses.filter(professor_id=professor_id)
    if status_filter == 'active':
        courses = courses.filter(status=True)
    elif status_filter == 'closed':
        courses = courses.filter(status=False)
    if class_type_filter:
        courses = courses.filter(class_type=class_type_filter)
    if course_level and course_level in ('1', '2', '3', '4', '5'):
        courses = courses.filter(course__iregex=r'^\D*' + course_level + r'\d{3}')
    return courses


@login_required
def export_schedule(request):
    if not request.user.is_superuser:
        messages.error(request, "Only admins can export the schedule.")
        return redirect('courses')
    term_filter = (request.GET.get('term') or '').strip()
    if not term_filter:
        messages.warning(request, "Select a term (use Add filter → Term) to export the schedule.")
        return redirect('courses')
    courses = _get_export_queryset(request).select_related('professor').prefetch_related('current_tas')
    headers = [
        'Term', 'Type', 'Course', 'Section', 'Course Title', 'Instructors', 'RoomName', 'TimeSlot',
        'Max Enroll', 'RoomSize',
        'Instructor Email', 'TAs Assigned', 'TAs Total', 'TA Names',
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = 'Schedule'
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
    for row_idx, course in enumerate(courses, 2):
        instructors_str = f"{course.instructor_last_name}, {course.instructor_first_name}" if (course.instructor_last_name or course.instructor_first_name) else (course.professor.get_full_name() if course.professor else '')
        if not instructors_str and course.professor:
            instructors_str = course.professor.get_full_name() or course.professor.email or ''
        ta_count = course.current_tas.count()
        ta_names = ', '.join(
            (ta.get_full_name() or ta.email or str(ta))
            for ta in course.current_tas.all()
        )
        instructor_email = (course.professor.email or '') if course.professor else ''
        ws.cell(row=row_idx, column=1, value=course.term)
        ws.cell(row=row_idx, column=2, value=course.class_type)
        ws.cell(row=row_idx, column=3, value=course.course)
        ws.cell(row=row_idx, column=4, value=course.section)
        ws.cell(row=row_idx, column=5, value=course.course_title)
        ws.cell(row=row_idx, column=6, value=instructors_str)
        ws.cell(row=row_idx, column=7, value=course.room_name)
        ws.cell(row=row_idx, column=8, value=course.timeslot)
        ws.cell(row=row_idx, column=9, value=course.max_enroll)
        ws.cell(row=row_idx, column=10, value=course.room_size)
        ws.cell(row=row_idx, column=11, value=instructor_email)
        ws.cell(row=row_idx, column=12, value=ta_count)
        ws.cell(row=row_idx, column=13, value=course.num_tas)
        ws.cell(row=row_idx, column=14, value=ta_names)
    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ''))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    safe_term = "".join(c if c.isalnum() or c in ' -' else '_' for c in term_filter)
    export_date = date.today().strftime('%Y-%m-%d')
    filename = f'Schedule_Export_{safe_term}_{export_date}.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

# Statuses that count toward the 5-application-per-term limit (Rejected/Withdrawn do not count)
APPLICATION_LIMIT_COUNTED_STATUSES = [
    ApplicationStatus.PENDING.value,
    ApplicationStatus.ACCEPTED.value,
    ApplicationStatus.CONFIRMED.value,
]


@login_required
def apply_to_course_v2(request, course_id):
    course = get_object_or_404(Course, id=course_id)

    # Closed courses do not accept applications
    if not course.status:
        messages.error(request, "This course is closed and is not accepting applications.")
        return redirect('courses')

    # Permission checks
    if request.user.is_professor:
        messages.error(request, "Professors cannot apply for TA positions.")
        return redirect('courses')

    if hasattr(request.user, 'is_ta') and request.user.is_ta():
        messages.error(request, "You are already a TA for a course.")
        return redirect('courses')

    # Check if already applied to this course
    if Application.objects.filter(student=request.user, course=course).exists():
        messages.warning(request, "You have already applied to this course.")
        return redirect('courses')

    # Require Eagle ID + resume before applying
    profile, _ = StudentProfile.objects.get_or_create(user=request.user)
    if not request.user.has_complete_profile_for_apply():
        messages.error(
            request,
                "Complete your profile before applying: add your 8-digit Eagle ID, Graduation Year, and upload a resume on your Profile page.",
        )
        return redirect("student_profile")

    def count_applications_for_term(student, term):
        if not term:
            return 0
        norm_term = term.strip()
        return Application.objects.filter(
            student=student,
            course__term__iexact=norm_term,
            status__in=APPLICATION_LIMIT_COUNTED_STATUSES,
        ).count()

    # Enforce 5-application limit per term: block when they already have 5
    current_term_count = count_applications_for_term(request.user, course.term)
    if current_term_count >= 5:
        if request.method == 'POST':
            return HttpResponse(
                "You have reached the 5-course application limit for this term.",
                status=400,
            )
        messages.error(request, "You have reached the 5-course application limit for this term.")
        return redirect('courses')

    custom_questions = list(course.custom_questions.all())

    if request.method == 'POST':
        form = ApplicationForm(request.POST)
        # Validate required custom question answers
        custom_errors = {}
        for question in custom_questions:
            if question.is_required and not (request.POST.get(f'question_{question.id}') or '').strip():
                custom_errors[question.id] = "This field is required."
        if form.is_valid() and not custom_errors:
            if request.user.is_ta():
                messages.error(request, "You are already a TA for a course.")
                return redirect('courses')
            # Re-check limit right before save
            if count_applications_for_term(request.user, course.term) >= 5:
                return HttpResponse(
                    "You have reached the 5-course application limit for this term.",
                    status=400,
                )
            app = form.save(commit=False)
            app.student = request.user
            app.course = course
            app.status = ApplicationStatus.PENDING.value
            # Snapshot profile at time of application
            app.skills_snapshot = [{"name": s.name} for s in profile.skills.all()]
            app.skills_additional_snapshot = (profile.skills_additional or "").strip()[:500]
            app.courses_snapshot = [
                {"course_name": pc.course_name}
                for pc in request.user.past_courses.all()
            ]
            app.save()
            # Save answers to custom course questions
            for question in course.custom_questions.all():
                answer_text = (request.POST.get(f'question_{question.id}') or '').strip()
                ApplicationAnswer.objects.create(
                    application=app,
                    question=question,
                    answer_text=answer_text,
                )
            # Copy resume and profile photo to application snapshots (what was submitted)
            import os

            if profile.resume:
                app.resume.save(os.path.basename(profile.resume.name), profile.resume, save=True)
            if profile.profile_photo:
                app.profile_photo.save(
                    os.path.basename(profile.profile_photo.name),
                    profile.profile_photo,
                    save=True,
                )

            messages.success(request, f"Successfully applied to {course.course}.")
            if course.professor:
                create_notification(
                    user=course.professor,
                    title=f"New application from {request.user.get_full_name() or request.user.email} for {course.course}",
                    target_url=reverse("application_detail", kwargs={"application_id": app.id}),
                )
            return redirect('courses')
    else:
        form = ApplicationForm()

    # Profile summary for template (resume link, skills, past courses)
    profile_skills = list(profile.skills.all().order_by('name'))
    profile_past_courses = list(request.user.past_courses.all().order_by('course_name'))
    posted = request.POST if request.method == 'POST' else {}
    errors = custom_errors if request.method == 'POST' else {}
    custom_questions_prefilled = [
        {
            'question': q,
            'answer_text': posted.get(f'question_{q.id}', ''),
            'error': errors.get(q.id, ''),
        }
        for q in custom_questions
    ]
    return render(request, 'application_form.html', {
        'form': form,
        'course': course,
        'profile': profile,
        'profile_skills': profile_skills,
        'profile_past_courses': profile_past_courses,
        'custom_questions_prefilled': custom_questions_prefilled,
    })


@login_required
def make_offer_v2(request, application_id):
    if request.method != "POST":
        return redirect("applications")

    if not (request.user.is_professor or request.user.is_superuser):
        messages.error(request, "Only professors or admins can make offers.")
        return redirect("applications")

    app = get_object_or_404(
        Application.objects.select_related("course", "student"),
        id=application_id,
    )
    course = app.course

    if request.user.is_professor and not request.user.is_superuser:
        if course.professor_id != request.user.id:
            messages.error(request, "You can only make offers for your own courses.")
            return redirect("applications")

    if app.status != ApplicationStatus.PENDING.value:
        messages.error(request, "You can only make an offer for a pending application.")
        return redirect("applications")

    if Offer.objects.filter(application=app).exists():
        messages.error(request, "An offer has already been sent for this application.")
        return redirect("applications")

    if not course.has_ta_offer_capacity():
        messages.error(
            request,
            "You cannot send more offers than there are TA slots for this course. "
            "Wait for responses or for a student to decline before offering another applicant.",
        )
        return redirect("applications")

    with transaction.atomic():
        app_locked = Application.objects.select_for_update().select_related("student").get(pk=app.pk)
        if app_locked.status != ApplicationStatus.PENDING.value:
            messages.error(request, "You can only make an offer for a pending application.")
            return redirect("applications")
        if Offer.objects.filter(application=app_locked).exists():
            messages.error(request, "An offer has already been sent for this application.")
            return redirect("applications")
        course_locked = Course.objects.select_for_update().get(pk=app_locked.course_id)
        used = course_locked.current_tas.count() + Offer.objects.filter(
            course=course_locked, status=OfferStatus.PENDING.value
        ).count()
        if not course_locked.num_tas or used >= course_locked.num_tas:
            messages.error(
                request,
                "You cannot send more offers than there are TA slots for this course. "
                "Wait for responses or for a student to decline before offering another applicant.",
            )
            return redirect("applications")
        Offer.objects.create(
            recipient=app_locked.student,
            course=course_locked,
            sender=request.user,
            application=app_locked,
            status=OfferStatus.PENDING.value,
        )
        app_locked.status = ApplicationStatus.ACCEPTED.value
        app_locked.save(update_fields=["status"])

    messages.success(request, f"Offer sent to {app.student.get_full_name()} for {course.course}.")
    create_notification(
        user=app.student,
        title=f"You received a TA offer for {course.course}",
        target_url=reverse("offers"),
    )

    if app.student.email:
        respond_url = app_site_absolute_url(reverse("offers"))
        send_notification_email(
            subject=f"TA Offer for {course.course}",
            recipients=app.student.email,
            message_lines=[
                f"Dear {app.student.get_full_name()},",
                f"Congratulations! You have received a TA offer for {course.course} — {course.course_title}.",
                "Open the link below to log in to TA Connect and accept or decline this offer:",
                respond_url,
            ],
        )

    return redirect("applications")

@login_required
def reject_application_v2(request, application_id):
    if not (request.user.is_professor or request.user.is_superuser):
        messages.error(request, "Only professors or admins can reject applications.")
        return redirect('applications')

    app = get_object_or_404(Application, id=application_id)

    # Step 1: Show rejection form (serves as the warning/confirmation)
    if request.method == "GET":
        return render(request, "application_reject.html", {"app": app})

    # Step 2: Submit rejection + optional feedback
    if request.method == "POST":
        rejection_feedback = (request.POST.get("rejection_feedback") or "").strip()

        app.reject(rejection_feedback)

        messages.success(request, f"Application for {app.student.get_full_name()} has been rejected.")
        create_notification(
            user=app.student,
            title=f"Your application for {app.course.course} was rejected",
            target_url=reverse("application_detail", kwargs={"application_id": app.id}),
        )

        if app.student.email:
            message_lines = [
                f"Dear {app.student.get_full_name()},",
                f"We regret to inform you that your TA application for {app.course.course} — {app.course.course_title} has been rejected.",
            ]
            if rejection_feedback:
                message_lines.extend([
                    "Feedback from the instructor:",
                    rejection_feedback,
                ])
            message_lines.append("You may browse and apply to other open courses on TA Connect.")
            send_notification_email(
                subject=f"Application Update for {app.course.course}",
                recipients=app.student.email,
                message_lines=message_lines,
            )

        return redirect("application_detail", application_id=app.id)

    return redirect("application_detail", application_id=app.id)

@login_required
def withdraw_application_v2(request, application_id):
    if request.method != 'POST':
        return redirect('application_detail', application_id=application_id)
    app = get_object_or_404(Application, id=application_id)
    if request.user != app.student:
        messages.error(request, "You are not authorized to withdraw this application.")
        return redirect('applications')
    if app.status not in (ApplicationStatus.PENDING.value, ApplicationStatus.ACCEPTED.value):
        messages.error(request, "Only pending or accepted applications can be withdrawn.")
        return redirect('applications')
    app.withdraw("Withdrawn by student")
    messages.success(request, "Your application has been withdrawn.")
    if app.course.professor:
        create_notification(
            user=app.course.professor,
            title=f"{app.student.get_full_name() or app.student.email} withdrew from {app.course.course}",
            target_url=reverse("application_detail", kwargs={"application_id": app.id}),
        )

    # Email notification to professor
    if app.course.professor and app.course.professor.email:
        send_notification_email(
            subject=f"Application Withdrawn for {app.course.course}",
            recipients=app.course.professor.email,
            message_lines=[
                f"Dear {app.course.professor.get_full_name()},",
                f"{app.student.get_full_name()} has withdrawn their TA application for {app.course.course} — {app.course.course_title}.",
            ],
        )

    return redirect('applications')


@login_required
def edit_application_v2(request, application_id):
    """Allow students to edit their own application (additional_information) when PENDING or ACCEPTED."""
    app = get_object_or_404(Application, id=application_id)
    if request.user != app.student:
        messages.error(request, "You are not authorized to edit this application.")
        return redirect('applications')
    if app.status not in (ApplicationStatus.PENDING.value, ApplicationStatus.ACCEPTED.value):
        messages.error(request, "Only pending or accepted applications can be edited.")
        return redirect('application_detail', application_id=app.id)

    course = app.course
    custom_questions = list(course.custom_questions.all())
    existing_answers = {a.question_id: a for a in app.custom_answers.select_related('question').all()}

    if request.method == 'POST':
        form = ApplicationForm(request.POST, instance=app)
        custom_errors = {}
        for question in custom_questions:
            if question.is_required and not (request.POST.get(f'question_{question.id}') or '').strip():
                custom_errors[question.id] = "This field is required."
        if form.is_valid() and not custom_errors:
            form.save()
            # Update or create answers for each custom question
            for question in custom_questions:
                answer_text = (request.POST.get(f'question_{question.id}') or '').strip()
                if question.id in existing_answers:
                    ans = existing_answers[question.id]
                    ans.answer_text = answer_text
                    ans.save(update_fields=['answer_text'])
                else:
                    ApplicationAnswer.objects.create(
                        application=app, question=question, answer_text=answer_text
                    )
            profile, _ = StudentProfile.objects.get_or_create(user=request.user)
            import os

            if profile.profile_photo:
                app.profile_photo.save(
                    os.path.basename(profile.profile_photo.name),
                    profile.profile_photo,
                    save=True,
                )
            elif app.profile_photo:
                app.profile_photo.delete(save=True)
            app.skills_snapshot = [{"name": s.name} for s in profile.skills.all()]
            app.skills_additional_snapshot = (profile.skills_additional or "").strip()[:500]
            app.save(update_fields=["skills_snapshot", "skills_additional_snapshot"])
            messages.success(request, "Your application has been updated.")
            return redirect('application_detail', application_id=app.id)
    else:
        form = ApplicationForm(instance=app)

    posted = request.POST if request.method == 'POST' else {}
    errors = locals().get('custom_errors', {})
    custom_questions_prefilled = [
        {
            'question': q,
            'answer_text': posted.get(f'question_{q.id}', existing_answers.get(q.id).answer_text if q.id in existing_answers else ''),
            'error': errors.get(q.id, ''),
        }
        for q in custom_questions
    ]

    return render(request, 'application_form.html', {
        'form': form,
        'course': course,
        'application': app,
        'is_edit': True,
        'custom_questions_prefilled': custom_questions_prefilled,
    })


@login_required
def accept_offer_v2(request, offer_id):
    if request.method == 'POST':
        offer = get_object_or_404(Offer, id=offer_id)
        if request.user != offer.recipient:
            messages.error(request, "You are not authorized to accept this offer.")
            return redirect('offers')
        # Exactly one TA assignment: block a second offer if already assigned elsewhere
        if offer.recipient.is_ta() and not offer.recipient.course_working_for.filter(
            pk=offer.course.pk
        ).exists():
            messages.error(
                request,
                "You are already a TA for another course. You can only hold one TA position at a time.",
            )
            return redirect('offers')
        with transaction.atomic():
            offer.accept()
            # Withdraw all other active applications (any term)
            Application.objects.filter(
                student=offer.recipient,
                status__in=[
                    ApplicationStatus.PENDING.value,
                    ApplicationStatus.ACCEPTED.value,
                ],
            ).exclude(id=offer.application_id).update(
                status=ApplicationStatus.WITHDRAWN.value,
                withdrawal_reason="Accepted a TA offer",
            )
            # Reject all other pending offers
            Offer.objects.filter(
                recipient=offer.recipient,
                status=OfferStatus.PENDING.value,
            ).exclude(id=offer.id).update(status=OfferStatus.REJECTED.value)
        messages.success(request, f"Congratulations! You are now a TA for {offer.course.course}.")
        create_notification(
            user=offer.recipient,
            title=f"You accepted the offer for {offer.course.course}",
            target_url=reverse("employment_onboarding"),
        )
        send_student_ta_acceptance_onboarding_email(
            offer.recipient, offer.course.course, offer.course.course_title
        )
        if offer.course.professor:
            create_notification(
                user=offer.course.professor,
                title=f"{offer.recipient.get_full_name() or offer.recipient.email} accepted your offer for {offer.course.course}",
                target_url=reverse("offers"),
            )

        # Email notification to professor
        if offer.course.professor and offer.course.professor.email:
            send_notification_email(
                subject=f"TA Offer Accepted for {offer.course.course}",
                recipients=offer.course.professor.email,
                message_lines=[
                    f"Dear {offer.course.professor.get_full_name()},",
                    f"{offer.recipient.get_full_name()} has accepted your TA offer for {offer.course.course} — {offer.course.course_title}.",
                    "They are now assigned as a TA for this course.",
                ],
            )

    return redirect('offers')

@login_required
def decline_offer_v2(request, offer_id):
    if request.method == 'POST':
        offer = get_object_or_404(Offer, id=offer_id)
        if request.user != offer.recipient:
            messages.error(request, "You are not authorized to decline this offer.")
            return redirect('offers')
        offer.reject()
        messages.info(request, f"You have declined the offer for {offer.course.course}.")
        if offer.course.professor:
            create_notification(
                user=offer.course.professor,
                title=f"{offer.recipient.get_full_name() or offer.recipient.email} declined your offer for {offer.course.course}",
                target_url=reverse("offers"),
            )

        # Email notification to professor
        if offer.course.professor and offer.course.professor.email:
            send_notification_email(
                subject=f"TA Offer Declined for {offer.course.course}",
                recipients=offer.course.professor.email,
                message_lines=[
                    f"Dear {offer.course.professor.get_full_name()},",
                    f"{offer.recipient.get_full_name()} has declined your TA offer for {offer.course.course} — {offer.course.course_title}.",
                    "You may consider making offers to other applicants.",
                ],
            )

    return redirect('offers')

@login_required
def remove_ta_v2(request, course_id, user_id):
    if request.method != 'POST':
        return redirect('course_overview', course_id=course_id)

    if not (request.user.is_professor or request.user.is_superuser):
        messages.error(request, "Only professors or admins can remove TAs.")
        return redirect('course_overview', course_id=course_id)

    course = get_object_or_404(Course, id=course_id)
    if request.user.is_professor and not request.user.is_superuser and course.professor_id != request.user.id:
        messages.error(request, "You can only manage your own courses.")
        return redirect('course_overview', course_id=course_id)

    ta = get_object_or_404(User, id=user_id)

    if ta not in course.current_tas.all():
        messages.warning(request, f"{ta.get_full_name()} is not a TA for this course.")
        return redirect('course_overview', course_id=course_id)

    with transaction.atomic():
        # Remove from course
        course.current_tas.remove(ta)
        # Re-open the course since there's now an open slot
        if course.current_tas.count() < course.num_tas:
            course.status = True
            course.save(update_fields=['status'])
        # Revert the offer to rejected
        Offer.objects.filter(
            recipient=ta, course=course, status=OfferStatus.ACCEPTED.value
        ).update(status=OfferStatus.REJECTED.value)
        # Revert the application to rejected
        Application.objects.filter(
            student=ta, course=course, status=ApplicationStatus.CONFIRMED.value
        ).update(status=ApplicationStatus.REJECTED.value)
        # Clear the student's assigned course
        ta.course_working_for.remove(course)

    messages.success(request, f"{ta.get_full_name()} has been removed as a TA for {course.course}.")
    create_notification(
        user=ta,
        title=f"You were removed as TA for {course.course}",
        target_url=reverse("offers"),
    )

    # Email notification to removed TA
    if ta.email:
        send_notification_email(
            subject=f"TA Assignment Update for {course.course}",
            recipients=ta.email,
            message_lines=[
                f"Dear {ta.get_full_name()},",
                f"You have been removed as a TA for {course.course} — {course.course_title}.",
                "If you have questions, please contact your course instructor.",
            ],
        )

    return redirect('course_overview', course_id=course_id)


@login_required
def delete_course_v2(request, course_id):
    if request.method != 'POST':
        return redirect('course_overview', course_id=course_id)

    if not request.user.is_superuser:
        messages.error(request, "Only admins can delete courses.")
        return redirect('courses')

    course = get_object_or_404(Course, id=course_id)
    course_name = f"{course.course} — {course.course_title}"

    with transaction.atomic():
        # Remove all assigned TAs
        for ta in course.current_tas.all():
            ta.course_working_for.remove(course)
        course.current_tas.clear()
        # Delete related offers and applications
        Offer.objects.filter(course=course).delete()
        Application.objects.filter(course=course).delete()
        course.delete()

    messages.success(request, f"Course {course_name} has been deleted.")
    return redirect('courses')


def _review_queue_nav_for_reviewer(application, user):
    """
    Same order as the Applications list: last name, first name, course, section, id.
    Superuser: all applications. Professor: only applications for their courses.
    Returns (prev_id, next_id, position_1_based, total) or (None, None, None, None) if N/A.
    """
    if user.is_superuser:
        qs = Application.objects.all()
    elif getattr(user, "is_professor", False) and application.course.professor_id == user.id:
        qs = Application.objects.filter(course__professor=user)
    else:
        return None, None, None, None
    qs = qs.select_related("student", "course").order_by(*APPLICATIONS_SORT_ORDER)
    ids = list(qs.values_list("id", flat=True))
    total = len(ids)
    try:
        idx = ids.index(application.id)
    except ValueError:
        return None, None, None, None
    prev_id = ids[idx - 1] if idx > 0 else None
    next_id = ids[idx + 1] if idx + 1 < total else None
    position = idx + 1
    return prev_id, next_id, position, total


@login_required
def application_detail_v2(request, application_id):
    app = get_object_or_404(
        Application.objects.select_related("student", "course", "course__professor"),
        id=application_id,
    )

    # Permission Check
    is_student_owner = request.user == app.student
    is_course_professor = request.user == app.course.professor

    if not (request.user.is_superuser or is_student_owner or is_course_professor):
        messages.error(request, "You are not authorized to view this application.")
        return redirect("dashboard")

    review_prev_app_id = review_next_app_id = None
    review_queue_position = review_queue_total = None
    show_applicant_navigation = request.user.is_superuser or (
        is_course_professor and getattr(request.user, "is_professor", False)
    )
    if show_applicant_navigation:
        (
            review_prev_app_id,
            review_next_app_id,
            review_queue_position,
            review_queue_total,
        ) = _review_queue_nav_for_reviewer(app, request.user)

    custom_answers = list(
        app.custom_answers.select_related('question').order_by('question__order', 'question__id')
    )

    return render(
        request,
        "application_detail.html",
        {
            "app": app,
            "show_applicant_navigation": show_applicant_navigation,
            "review_prev_app_id": review_prev_app_id,
            "review_next_app_id": review_next_app_id,
            "review_queue_position": review_queue_position,
            "review_queue_total": review_queue_total,
            "custom_answers": custom_answers,
        },
    )


@login_required
def serve_application_resume(request, application_id):
    """Serve the resume snapshot for an application (professor or superuser only)."""
    app = get_object_or_404(Application, id=application_id)
    if not (request.user.is_superuser or request.user == app.course.professor):
        messages.error(request, "You are not authorized to view this resume.")
        return redirect('dashboard')
    if not app.resume:
        messages.warning(request, "No resume was submitted with this application.")
        return redirect('application_detail', application_id=application_id)
    try:
        file_handle = app.resume.open('rb')
        filename = app.resume.name.split('/')[-1] if app.resume.name else 'resume'
        if filename.lower().endswith('.pdf'):
            content_type = 'application/pdf'
        elif filename.lower().endswith('.doc'):
            content_type = 'application/msword'
        elif filename.lower().endswith('.docx'):
            content_type = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        else:
            content_type = 'application/octet-stream'
        response = FileResponse(file_handle, content_type=content_type)
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response
    except (ValueError, OSError):
        messages.error(request, "Resume file could not be found.")
        return redirect('application_detail', application_id=application_id)


@login_required
def serve_application_applicant_photo(request, application_id):
    """Serve the profile photo snapshot for an application (student owner, course professor, or admin)."""
    app = get_object_or_404(Application, id=application_id)
    if not (
        request.user.is_superuser
        or request.user == app.student
        or request.user == app.course.professor
    ):
        return HttpResponse(status=403)
    if not app.profile_photo:
        return HttpResponse(status=404)
    try:
        file_handle = app.profile_photo.open("rb")
        filename = (app.profile_photo.name or "").lower()
        if filename.endswith(".png"):
            content_type = "image/png"
        elif filename.endswith(".gif"):
            content_type = "image/gif"
        elif filename.endswith(".webp"):
            content_type = "image/webp"
        else:
            content_type = "image/jpeg"
        response = FileResponse(file_handle, content_type=content_type)
        response["Cache-Control"] = "private, max-age=3600"
        return response
    except (ValueError, OSError):
        return HttpResponse(status=404)
